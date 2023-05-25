from openpyxl import Workbook
import openpyxl as op


condi_VUC = [] # Состояние ВУЦ
condi_student = [] # Состояние студента
course = [] # Курс
surname = [] # Фамилия
name = [] # Имя
lastname = [] # Отчество
birth = [] # Дата рождения
facul = [] # Факультет
sex = [] # Пол
nabor = [] # Набор
code_spec = [] # Код специальности
group= [] # Группа
name_spec = [] # Наименование специальности
country = [] # Гражданство
phone = [] # Телефон
form_edu = [] # Форма обучения
name_fac = [] # Наименование факультета
familu_status = [] # Семейное положение
birth_place = [] # Место рождения
email = [] # E-mail
ser_passport = [] # Серия паспорта
number_passport = [] # Номер паспорта
issue_date = [] # Дата выдачи
kem_vidan = [] # Кем выдан
cvalification = [] # Квалификация
name_cvalification = [] # Наименование квалификации
edu_time = [] # Срок обучения
number_studak = [] # Номер студенческого билета
number_PF = [] # Номер ПФ
number_zachetki = [] # Номер зачетной книжки
edu_status=[] # Статус обучающегося
est_release_date = [] # Предполагаемая дата выпуска
passport = [] # Паспорт

filename = 'Из Паруса сведения.xlsx' #Указываем файл, который хотим читать.
#Сначала проверить, что у файла excel расширение .xlsx. Если нет, то копировать в новый файл в Excel 2007 и новее/
#Файлы нужно помещать в корень, но в репозе на гите их нет дабы не произошло утечки данных. Просто закиньте в корень.
val_dict={}
wb = op.load_workbook(filename, data_only = True)
sheet = wb.active

max_rows = sheet.max_row
max_cols = sheet.max_column



for i in range(2, max_rows+1):
    condi_VUC.append(sheet.cell(row=i, column=1).value)
    condi_student.append(sheet.cell(row=i, column=2).value)
    course.append(sheet.cell(row=i, column=3).value)
    surname.append(sheet.cell(row=i, column=4).value)
    name.append(sheet.cell(row=i, column=5).value)
    lastname.append(sheet.cell(row=i, column=6).value)
    birth.append(sheet.cell(row=i, column=7).value)
    facul.append(sheet.cell(row=i, column=8).value)
    sex.append(sheet.cell(row=i, column=9).value)
    nabor.append(sheet.cell(row=i, column=10).value)
    code_spec.append(sheet.cell(row=i, column=11).value)
    group.append(sheet.cell(row=i, column=12).value)
    name_spec.append(sheet.cell(row=i, column=13).value)
    country.append(sheet.cell(row=i, column=14).value)
    phone.append(sheet.cell(row=i, column=15).value)
    form_edu.append(sheet.cell(row=i, column=16).value)
    name_fac.append(sheet.cell(row=i, column=17).value)
    familu_status.append(sheet.cell(row=i, column=18).value)
    birth_place.append(sheet.cell(row=i, column=19).value)
    email.append(sheet.cell(row=i, column=20).value)
    ser_passport.append(sheet.cell(row=i, column=21).value)
    number_passport.append(sheet.cell(row=i, column=22).value)
    issue_date.append(sheet.cell(row=i, column=23).value)
    kem_vidan.append(sheet.cell(row=i, column=24).value)
    cvalification.append(sheet.cell(row=i, column=25).value)
    name_cvalification.append(sheet.cell(row=i, column=26).value)
    edu_time.append(sheet.cell(row=i, column=27).value)
    number_studak.append(sheet.cell(row=i, column=28).value)
    number_PF.append(sheet.cell(row=i, column=29).value)
    number_zachetki.append(sheet.cell(row=i, column=30).value)
    edu_status.append(sheet.cell(row=i, column=31).value)
    est_release_date.append(sheet.cell(row=i, column=32).value)
    passport.append(sheet.cell(row=i, column=33).value)




wb = Workbook()
ws = wb.active

#Код, который выводит в первый столбец инфу с первого массива. Нужно сделать +- то же самое для остальных и все по сути.

ws['A1'] = 'Состояние ВУЦ'
for i in range(2, len(condi_VUC)):
    if i == 0:
        i = 1
    temp = 'A'+str(i)
    ws[temp] = condi_VUC[i-2]

wb.save("itog.xlsx")